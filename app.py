import streamlit as st
import zipfile
import io
import re
import difflib
import time
from datetime import date
from typing import List, Dict
from anonymizer import MQXLIFFAnonymizer, load_dictionary_terms
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

st.set_page_config(
    page_title="Anonymizer",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded"
)

CUSTOM_CSS = """
<style>
    .main {
        background-color: #ffffff !important;
    }
    .stApp {
        background-color: #ffffff !important;
    }
    .stApp, .stApp p, .stApp span, .stApp label, .stApp div {
        color: #130e45 !important;
    }
    
    /* Global scrollbar styles */
    *::-webkit-scrollbar {
        width: 16px !important;
        height: 16px !important;
    }
    *::-webkit-scrollbar-track {
        background: #d0d3d4 !important;
        border-radius: 8px !important;
    }
    *::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0) !important;
        border-radius: 8px !important;
        border: 2px solid #d0d3d4 !important;
    }
    *::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488) !important;
    }
    
    /* Firefox scrollbar */
    * {
        scrollbar-width: auto;
        scrollbar-color: #1a5488 #d0d3d4;
    }
    [data-testid="stSidebar"] {
        background-color: #e0e3e4 !important;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] span, [data-testid="stSidebar"] label {
        color: #130e45 !important;
    }
    h1, h2, h3 {
        color: #1a5488 !important;
    }
    .stButton > button,
    [data-testid="stBaseButton-primary"] {
        background-color: #0e7bc0 !important;
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 800;
    }
    .stButton > button p,
    .stButton > button span,
    .stButton > button div,
    [data-testid="stBaseButton-primary"] p,
    [data-testid="stBaseButton-primary"] span {
        color: white !important;
        background-color: transparent !important;
        padding: 0 !important;
    }
    [data-testid="stBaseButton-primary"] {
        font-weight: 900 !important;
        font-size: 1.1rem !important;
    }
    .stButton > button:hover,
    [data-testid="stBaseButton-primary"]:hover {
        background-color: #134277 !important;
    }
    .stat-card {
        background-color: #e0e3e4;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        border-left: 4px solid #1a5488;
    }
    .stat-card-safe-regex { border-left-color: #0e7bc0; }
    .stat-card-regex-ct { border-left-color: #6f42c1; }
    .stat-card-presidio { border-left-color: #e67e22; }
    .stat-card-biomedical { border-left-color: #e74c3c; }
    .stat-card-proper-names { border-left-color: #28a745; }
    .stat-card-dictionary { border-left-color: #17a2b8; }
    .stat-number {
        font-size: 2rem;
        font-weight: bold;
        color: #130e45;
    }
    .stat-label {
        color: #5e5f6b;
        font-size: 0.9rem;
    }
    .sidebar-divider {
        border: none;
        border-top: 1px solid #c0c3c4;
        margin: 0.8rem 0;
    }
    .app-footer {
        text-align: center;
        padding: 1.5rem 0 0.5rem 0;
        color: #8a8b96 !important;
        font-size: 0.8rem;
        border-top: 1px solid #e0e3e4;
        margin-top: 2rem;
    }
    .app-footer p, .app-footer span {
        color: #8a8b96 !important;
    }
    .preview-box {
        background-color: #f8f9fa;
        border: 1px solid #bcbdbe;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .preview-container {
        max-height: 500px;
        overflow-y: auto;
        padding-right: 10px;
        margin: 1rem 0;
        border: 1px solid #e0e3e4;
        border-radius: 8px;
        background-color: #fafafa;
    }
    .preview-container::-webkit-scrollbar {
        width: 18px;
    }
    .preview-container::-webkit-scrollbar-track {
        background: #d0d3d4;
        border-radius: 9px;
        box-shadow: inset 0 0 3px rgba(0,0,0,0.2);
    }
    .preview-container::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 9px;
        border: 3px solid #d0d3d4;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    .preview-container::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    div[data-testid="stExpander"] {
        max-height: 600px;
        overflow-y: auto;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar {
        width: 20px;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-track {
        background: #c8cbcc;
        border-radius: 10px;
        box-shadow: inset 0 0 4px rgba(0,0,0,0.25);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 10px;
        border: 3px solid #c8cbcc;
        box-shadow: 0 2px 5px rgba(0,0,0,0.35);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    .exclude-badge {
        background-color: #dc3545;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 4px;
        font-size: 0.75rem;
        font-weight: bold;
        margin-left: 0.5rem;
    }
    .excluded-segment {
        border: 2px dashed #dc3545 !important;
        background-color: #fff0f0 !important;
    }
    .before-text {
        color: #dc3545;
        background-color: #ffe6e6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .after-text {
        color: #28a745;
        background-color: #e6ffe6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .section-header {
        background-color: #7cb4db;
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #e0e3e4;
        border-left: 4px solid #0e7bc0;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    div[data-testid="stFileUploader"] {
        background-color: #f8f9fa;
        border: 2px dashed #7cb4db;
        border-radius: 10px;
        padding: 1rem;
    }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def count_words(text: str) -> int:
    """Count words in text, ignoring XML tags and whitespace."""
    if not text:
        return 0
    import re
    clean = re.sub(r'<[^>]+>', ' ', text)
    clean = re.sub(r'\s+', ' ', clean).strip()
    words = [w for w in clean.split() if len(w) > 0]
    return len(words)


def segment_word_count(preview: dict) -> int:
    """Returns the max word count between source and target."""
    source_words = count_words(preview.get('source_before', ''))
    target_words = count_words(preview.get('target_before', ''))
    return max(source_words, target_words)


def is_junk_segment(preview: dict, min_words_junk: int = 3) -> bool:
    """Detects junk/short original segments that pollute TM databases.
    Returns True if the segment should be excluded."""
    import re
    source = preview.get('source_before', '').strip()
    target = preview.get('target_before', '').strip()
    
    if not source and not target:
        return True
    
    for text in [source, target]:
        if not text:
            continue
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r'\s+', ' ', clean).strip()
        if not clean:
            continue
        if re.fullmatch(r'[\d\s\-.,;:!?¿¡()\[\]{}/\\|@#$%^&*+=<>~`"\'°ºª•–—…\u2022\u2013\u2014\u2026]+', clean):
            return True
        words = [w for w in clean.split() if len(w) > 0]
        if len(words) < min_words_junk:
            return True
    
    return False


def render_stat_card(label: str, value: int, col, css_class: str = ""):
    with col:
        st.markdown(f"""
        <div class="stat-card {css_class}">
            <div class="stat-number">{value}</div>
            <div class="stat-label">{label}</div>
        </div>
        """, unsafe_allow_html=True)


def _normalize_lang_code(code: str) -> str:
    """Normalize language code to standard format (e.g., de-de -> de-DE, en -> en)."""
    parts = code.strip().split("-")
    if len(parts) == 2:
        return f"{parts[0].lower()}-{parts[1].upper()}"
    return parts[0].lower()


def strip_inline_tags(text: str) -> str:
    if not text:
        return text
    cleaned = re.sub(r'<[^>]+>', '', text)
    cleaned = re.sub(r'\{/?(\d+)?\}', '', cleaned)
    cleaned = re.sub(r'  +', ' ', cleaned)
    return cleaned.strip()


def generate_clean_tmx(previews: dict, results: dict, originals: dict,
                       filter_junk: bool, min_words_junk: int,
                       filter_short: bool, min_words: int,
                       exclude_modified: bool, exclusion_threshold: float,
                       excluded_segments: dict, no_anon_segments: dict,
                       dedup_tmx: bool = True, dedup_threshold: int = 100) -> tuple:
    """Generate a clean TMX containing only valid anonymized segments."""
    from lxml import etree
    
    src_lang = "en"
    tgt_lang = "es"
    
    for filename in originals:
        try:
            tree = etree.fromstring(originals[filename])
            is_tmx = filename.lower().endswith(".tmx")
            if is_tmx:
                header = tree.find(".//header")
                if header is not None:
                    src_lang = header.get("srclang", "en")
                    tus = tree.xpath("//tu/tuv")
                    langs = set()
                    for tuv in tus:
                        lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                        if lang:
                            langs.add(lang)
                    for lang in langs:
                        if lang.lower() != src_lang.lower():
                            tgt_lang = lang
                            break
            else:
                nsmap = tree.nsmap
                default_ns = nsmap.get(None, '')
                if default_ns:
                    ns = {'x': default_ns}
                    file_els = tree.xpath('//x:file', namespaces=ns)
                else:
                    file_els = tree.xpath('//file')
                if file_els:
                    src_lang = file_els[0].get("source-language", "en")
                    tgt_lang = file_els[0].get("target-language", "es")
            break
        except Exception:
            pass
    
    src_lang = _normalize_lang_code(src_lang)
    tgt_lang = _normalize_lang_code(tgt_lang)
    
    tmx_root = etree.Element("tmx", version="1.4")
    header = etree.SubElement(tmx_root, "header",
                              creationtool="Anonymizer",
                              datatype="PlainText",
                              segtype="sentence",
                              srclang=src_lang)
    body = etree.SubElement(tmx_root, "body")
    
    valid_count = 0
    no_anon_skipped = 0
    dedup_count = 0
    dedup_details = []
    excluded_ids = []
    replacement_token = st.session_state.get('replacement_token', '███')
    seen_exact = set()
    seen_segments = []
    
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            
            if filter_junk and is_junk_segment(preview, min_words_junk):
                sk = f"skipjunk_{segment_key}"
                if not st.session_state.get(sk, False):
                    excluded_ids.append(preview['segment'])
                    continue
            
            if no_anon_segments.get(segment_key, False):
                no_anon_skipped += 1
                continue
            
            if filter_short and segment_word_count(preview) < min_words:
                excluded_ids.append(preview['segment'])
                continue
            
            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            
            token_esc = re.escape(replacement_token)
            consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)
            
            if not source_text and not target_text:
                excluded_ids.append(preview['segment'])
                continue
            
            if exclude_modified:
                should_exclude = excluded_segments.get(segment_key, None)
                if should_exclude is not False:
                    has_token_src = replacement_token in (preview.get('source_after', ''))
                    has_token_tgt = replacement_token in (preview.get('target_after', ''))
                    if has_token_src or has_token_tgt:
                        src_before = preview.get('source_before', '')
                        tgt_before = preview.get('target_before', '')
                        src_after = preview.get('source_after', '')
                        tgt_after = preview.get('target_after', '')
                        src_pct = (1 - len(src_after.replace(replacement_token, '')) / max(len(src_before), 1)) * 100 if src_before else 0
                        tgt_pct = (1 - len(tgt_after.replace(replacement_token, '')) / max(len(tgt_before), 1)) * 100 if tgt_before else 0
                        max_pct = max(src_pct, tgt_pct)
                        if max_pct >= exclusion_threshold:
                            excluded_ids.append(preview['segment'])
                            continue
            
            if dedup_tmx:
                seg_pair = (source_text, target_text)
                if dedup_threshold >= 100:
                    if seg_pair in seen_exact:
                        dedup_count += 1
                        dedup_details.append({
                            "file": filename, "segment": preview['segment'],
                            "source": source_text, "target": target_text,
                            "similarity": 100.0
                        })
                        continue
                    seen_exact.add(seg_pair)
                else:
                    combined = source_text + " ||| " + target_text
                    is_dup = False
                    match_ratio = 0.0
                    threshold_ratio = dedup_threshold / 100.0
                    for seen_combined in seen_segments:
                        ratio = difflib.SequenceMatcher(None, combined, seen_combined).ratio()
                        if ratio >= threshold_ratio:
                            is_dup = True
                            match_ratio = ratio * 100
                            break
                    if is_dup:
                        dedup_count += 1
                        dedup_details.append({
                            "file": filename, "segment": preview['segment'],
                            "source": source_text, "target": target_text,
                            "similarity": match_ratio
                        })
                        continue
                    seen_segments.append(combined)
            
            tu = etree.SubElement(body, "tu")
            
            tuv_src = etree.SubElement(tu, "tuv")
            tuv_src.set("{http://www.w3.org/XML/1998/namespace}lang", src_lang)
            seg_src = etree.SubElement(tuv_src, "seg")
            seg_src.text = source_text
            
            tuv_tgt = etree.SubElement(tu, "tuv")
            tuv_tgt.set("{http://www.w3.org/XML/1998/namespace}lang", tgt_lang)
            seg_tgt = etree.SubElement(tuv_tgt, "seg")
            seg_tgt.text = target_text
            
            valid_count += 1
    
    result = b'<?xml version="1.0" encoding="UTF-8"?>\n'
    result += etree.tostring(tmx_root, encoding="unicode", pretty_print=True).encode("utf-8")
    
    return result, valid_count, no_anon_skipped, excluded_ids, dedup_count, dedup_details


def generate_changes_excel(dedup_details: list = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anonymization Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a5488", end_color="1a5488", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1a5488")
    border = Border(
        left=Side(style='thin', color='bcbdbe'),
        right=Side(style='thin', color='bcbdbe'),
        top=Side(style='thin', color='bcbdbe'),
        bottom=Side(style='thin', color='bcbdbe')
    )
    alt_fill = PatternFill(start_color="e8f4fc", end_color="e8f4fc", fill_type="solid")
    total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    total_font = Font(bold=True, size=11)
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "ANONYMIZATION REPORT - MQXLIFF/TMX Anonymizer v6.2"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "STATISTICS SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    stats_headers = ["File", "Safe Regex", "Regex CT IDs", "Presidio", "ScispaCy", "Proper Names", "Dictionary", "Total"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    grand_total = 0
    for filename, stats in st.session_state.all_stats.items():
        file_total = stats.get("safe_regex", 0) + stats.get("regex_ct", 0) + stats.get("presidio_pii", 0) + stats.get("biomedical", 0) + stats.get("proper_names", 0) + stats.get("dictionary", 0)
        grand_total += file_total
        data = [filename, stats.get("safe_regex", 0), stats["regex_ct"], stats["presidio_pii"], stats["biomedical"], stats.get("proper_names", 0), stats["dictionary"], file_total]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
        row += 1
    
    total_data = ["GRAND TOTAL", "", "", "", "", "", "", grand_total]
    for col, value in enumerate(total_data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='center')
    
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "CHANGES DETAIL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    detail_headers = ["File", "Segment", "Type", "Original Text", "Anonymized Text"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    alt_row = False
    filter_junk = st.session_state.get('filter_junk', False)
    min_words_junk = st.session_state.get('min_words_junk', 2)
    filter_short = st.session_state.get('filter_short_segments', False)
    min_words = st.session_state.get('min_words', 5)
    
    for filename, file_previews in st.session_state.previews.items():
        for preview in file_previews:
            if filter_junk and is_junk_segment(preview, min_words_junk):
                continue
            if not preview.get('changed', True):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue
            
            if preview['source_before'] != preview['source_after']:
                data = [filename, preview['segment'], "Source", preview['source_before'], preview['source_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
            if preview['target_before'] != preview['target_after']:
                data = [filename, preview['segment'], "Target", preview['target_before'], preview['target_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    ws2 = wb.create_sheet(title="Excluded Segments")
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "EXCLUDED SEGMENTS - Anonymizer v6.2"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ex_row = 3
    
    short_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    tm_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    
    replacement_token = st.session_state.get('replacement_token', '███')
    threshold = st.session_state.get('exclusion_threshold', 20)
    exclude_enabled = st.session_state.get('exclude_modified_targets', False)
    excluded_segments = st.session_state.get('excluded_segments', {})
    
    has_junk = filter_junk
    has_short = filter_short
    has_tm = exclude_enabled
    
    junk_fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
    
    if has_junk:
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT SEGMENTS - Excluded from TM (<{min_words_junk} words or only numbers/symbols)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="495057")
        ws2[f'A{ex_row}'].fill = junk_fill
        ex_row += 1
        
        junk_headers = ["File", "Segment", "Source (original)", "Target (original)"]
        for col, header in enumerate(junk_headers, 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        
        junk_count = 0
        for filename, file_previews in st.session_state.previews.items():
            for preview in file_previews:
                if is_junk_segment(preview, min_words_junk):
                    seg_key = f"{filename}_{preview['segment']}"
                    if st.session_state.get(f"skipjunk_{seg_key}", False):
                        continue
                    junk_count += 1
                    data = [
                        filename,
                        preview['segment'],
                        preview['source_before'],
                        preview['target_before'],
                    ]
                    for col, value in enumerate(data, 1):
                        cell = ws2.cell(row=ex_row, column=col, value=value)
                        cell.border = border
                        cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ex_row += 1
        
        if junk_count == 0:
            ws2.cell(row=ex_row, column=1, value="No short segments found.").font = Font(italic=True, color="666666")
            ex_row += 1
        else:
            ws2.merge_cells(f'A{ex_row}:F{ex_row}')
            ws2[f'A{ex_row}'] = f"Total: {junk_count} short segments excluded from TM"
            ws2[f'A{ex_row}'].font = total_font
            ws2[f'A{ex_row}'].fill = junk_fill
            ex_row += 1
        
        ex_row += 2
    
    if has_short:
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT ANONYMIZED SEGMENTS - Excluded from TM (less than {min_words} words)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="856404")
        ws2[f'A{ex_row}'].fill = short_fill
        ex_row += 1
        
        short_headers = ["File", "Segment", "Source (original)", "Target (original)", "Source (anonymized)", "Target (anonymized)"]
        for col, header in enumerate(short_headers, 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        
        short_count = 0
        for filename, file_previews in st.session_state.previews.items():
            for preview in file_previews:
                if has_junk and is_junk_segment(preview, min_words_junk):
                    continue
                if not preview.get('changed', True):
                    continue
                if segment_word_count(preview) < min_words:
                    short_count += 1
                    data = [
                        filename,
                        preview['segment'],
                        preview['source_before'],
                        preview['target_before'],
                        preview['source_after'],
                        preview['target_after'],
                    ]
                    for col, value in enumerate(data, 1):
                        cell = ws2.cell(row=ex_row, column=col, value=value)
                        cell.border = border
                        cell.fill = PatternFill(start_color="fffbe6", end_color="fffbe6", fill_type="solid")
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ex_row += 1
        
        if short_count == 0:
            ws2.cell(row=ex_row, column=1, value="No short segments found.").font = Font(italic=True, color="666666")
            ex_row += 1
        else:
            ws2.merge_cells(f'A{ex_row}:F{ex_row}')
            ws2[f'A{ex_row}'] = f"Total: {short_count} short segments excluded from TM"
            ws2[f'A{ex_row}'].font = total_font
            ws2[f'A{ex_row}'].fill = short_fill
            ex_row += 1
        
        ex_row += 2
    
    if has_tm:
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"TM EXCLUDED SEGMENTS (redaction >= {threshold}%)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="721c24")
        ws2[f'A{ex_row}'].fill = tm_fill
        ex_row += 1
        
        tm_headers = ["File", "Segment", "Redaction %", "Source (anonymized)", "Target (anonymized)", "Reason"]
        for col, header in enumerate(tm_headers, 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        
        tm_count = 0
        for filename, file_previews in st.session_state.previews.items():
            for preview in file_previews:
                if not preview.get('changed', True):
                    continue
                has_token = (replacement_token in preview.get('source_after', '') or 
                            replacement_token in preview.get('target_after', ''))
                if not has_token:
                    continue
                if filter_junk and is_junk_segment(preview, min_words_junk):
                    continue
                if filter_short and segment_word_count(preview) < min_words:
                    continue
                
                segment_key = f"{filename}_{preview['segment']}"
                target_text = preview.get('target_after', '')
                target_pct = 0
                if target_text:
                    words = target_text.split()
                    if words:
                        target_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                
                is_excluded = excluded_segments.get(segment_key, True) if target_pct >= threshold else False
                
                if is_excluded:
                    tm_count += 1
                    data = [
                        filename,
                        preview['segment'],
                        f"{target_pct:.1f}%",
                        preview['source_after'],
                        preview['target_after'],
                        f"Redaction >= {threshold}%",
                    ]
                    for col, value in enumerate(data, 1):
                        cell = ws2.cell(row=ex_row, column=col, value=value)
                        cell.border = border
                        cell.fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ex_row += 1
        
        if tm_count == 0:
            ws2.cell(row=ex_row, column=1, value="No TM-excluded segments found.").font = Font(italic=True, color="666666")
            ex_row += 1
        else:
            ws2.merge_cells(f'A{ex_row}:F{ex_row}')
            ws2[f'A{ex_row}'] = f"Total: {tm_count} segments excluded from TM"
            ws2[f'A{ex_row}'].font = total_font
            ws2[f'A{ex_row}'].fill = tm_fill
            ex_row += 1
    
    dedup_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    has_dedup = dedup_details and len(dedup_details) > 0
    
    if has_dedup:
        ex_row += 2
        dedup_threshold_val = st.session_state.get('dedup_threshold', 100)
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"DUPLICATE SEGMENTS - Excluded from TMX (similarity >= {dedup_threshold_val}%)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="0c5460")
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
        
        dedup_headers = ["File", "Segment", "Similarity", "Source", "Target"]
        for col, header in enumerate(dedup_headers, 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        
        for dd in dedup_details:
            data = [
                dd['file'],
                dd['segment'],
                f"{dd['similarity']:.1f}%",
                dd['source'],
                dd['target'],
            ]
            for col, value in enumerate(data, 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="e8f6f8", end_color="e8f6f8", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(dedup_details)} duplicate segments excluded from TMX"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
    
    if not has_junk and not has_short and not has_tm and not has_dedup:
        ws2.cell(row=ex_row, column=1, value="No filters active. No segments were excluded.").font = Font(italic=True, color="666666")
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 45
    ws2.column_dimensions['F'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0e7bc0 0%, #1a5488 100%); 
                padding: 2rem; 
                border-radius: 12px; 
                margin-bottom: 2rem;
                box-shadow: 0 4px 15px rgba(26, 84, 136, 0.3);">
        <h1 style="color: white !important; margin: 0; font-size: 2.5rem; font-weight: 700; letter-spacing: 8px;">
            Anonymizer <span style="background: rgba(255,255,255,0.2); color: rgba(255,255,255,0.95); font-size: 0.9rem; font-weight: 500; padding: 3px 10px; border-radius: 20px; vertical-align: middle; letter-spacing: 1px; border: 1px solid rgba(255,255,255,0.3);">v6.2</span>
        </h1>
        <p style="color: rgba(255,255,255,0.95) !important; margin: 0.5rem 0 0 0; font-size: 1.1rem; font-weight: 400; letter-spacing: 1px;">
            Anonymize & clean bilingual memoQ documents
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        replacement_token = st.text_input("Replacement token", value="███")
        st.session_state['replacement_token'] = replacement_token
        process_source = True
        process_target = True
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Multilingual layers")
        use_safe_regex = st.checkbox("Safe Regex", value=True, help="Emails, phones, URLs, IDs, addresses, titled names, etc.")
        use_proper_names = st.checkbox("Proper Names", value=False, help="Structural person name detection (labels, initials, scoring)")
        use_dictionary = st.checkbox("Custom dictionary", value=True, help="Blacklist: manually loaded terms to anonymize")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### EN > ES layers")
        use_regex = st.checkbox("Clinical ID Regex", value=False, help="NCT IDs, EudraCT, Protocol IDs, Subject IDs, etc.")
        use_presidio = st.checkbox("Presidio", value=False, help="Emails, phone numbers, person names, addresses")
        use_biomedical = st.checkbox("ScispaCy", value=False, help="Drugs, pharmaceutical organizations")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Filters")
        dedup_tmx = st.checkbox("Deduplicate TMX segments", value=True, help="Remove duplicate or similar segments from the clean TMX")
        dedup_threshold = st.slider("Similarity threshold (%)", 50, 100, 100, 5, help="100% = exact match (recommended for large files). Lower values use fuzzy matching.") if dedup_tmx else 100
        st.session_state['dedup_tmx'] = dedup_tmx
        st.session_state['dedup_threshold'] = dedup_threshold
        filter_junk = st.checkbox("Exclude short segments", value=True, help="Removes short segments: less than min. words or only numbers/symbols")
        min_words_junk = st.slider("Minimum words (short segments)", 2, 10, 2, 1) if filter_junk else 2
        st.session_state['filter_junk'] = filter_junk
        st.session_state['min_words_junk'] = min_words_junk
        filter_short_segments = st.checkbox("Exclude short anon. segments", value=True, help="Excludes segments that are too short after anonymization")
        min_words = st.slider("Minimum words (anon. segments)", 2, 10, 5, 1) if filter_short_segments else 5
        st.session_state['filter_short_segments'] = filter_short_segments
        st.session_state['min_words'] = min_words
        exclude_modified_targets = st.checkbox("Exclude heavily anonymized", value=True)
        if exclude_modified_targets:
            exclusion_threshold = st.slider("Threshold (%)", 10, 90, 50, 5, help="Segments above X% anonymized excluded")
            exclude_source_too = True
        else:
            exclusion_threshold = 20
            exclude_source_too = False
        st.session_state['exclusion_threshold'] = exclusion_threshold
        st.session_state['exclude_source_too'] = exclude_source_too
        st.session_state['exclude_modified_targets'] = exclude_modified_targets
    
    tab1, tab2, tab3 = st.tabs(["📤 Upload", "📝 Preview", "📥 Download"])
    
    with tab1:
        st.markdown("### Upload MQXLIFF / TMX files")
        
        mqxliff_files = st.file_uploader(
            "Select one or more .mqxliff or .tmx files",
            type=["mqxliff", "tmx"],
            accept_multiple_files=True,
            help="You can upload multiple files for batch processing"
        )
        
        if mqxliff_files:
            st.success(f"✅ {len(mqxliff_files)} file(s) loaded")
            for f in mqxliff_files:
                ext = f.name.rsplit(".", 1)[-1].upper() if "." in f.name else "?"
                st.write(f"- {f.name} ({f.size / 1024:.1f} KB) — {ext}")
        
        st.markdown("---")
        st.markdown("### Custom dictionary (blacklist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will be anonymized:</strong><br>
            These terms will be forcefully anonymized even if not detected automatically.<br>
            One term per line or separated by commas:<br>
            <code>Example</code><br>
            <code>Example project, info@company.com</code><br>
            <code>Brand®, www.example.com</code>
        </div>
        """, unsafe_allow_html=True)
        
        dictionary_file = st.file_uploader(
            "Upload TXT file with sensitive terms (optional)",
            type=["txt"],
            help="Terms not well detected by Presidio or ScispaCy"
        )
        
        dictionary_terms = set()
        if dictionary_file:
            dict_raw = dictionary_file.read()
            if dict_raw[:3] == b'\xef\xbb\xbf':
                dict_raw = dict_raw[3:]
            elif dict_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                dict_raw = dict_raw.decode('utf-16').encode('utf-8')
            try:
                content = dict_raw.decode("utf-8")
            except UnicodeDecodeError:
                content = dict_raw.decode("latin-1")
            dictionary_terms = load_dictionary_terms(content)
            st.success(f"✅ {len(dictionary_terms)} unique terms loaded")
            
            with st.expander("View loaded terms"):
                for term in sorted(dictionary_terms):
                    st.write(f"- {term}")
        
        st.markdown("### Protected terms (whitelist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will NOT be anonymized:</strong><br>
            These terms will be preserved even if detected.<br>
            One term per line or separated by commas.
        </div>
        """, unsafe_allow_html=True)
        
        whitelist_file = st.file_uploader(
            "Upload TXT file with protected terms (optional)",
            type=["txt"],
            help="Terms that should never be anonymized",
            key="whitelist_uploader"
        )
        
        whitelist_terms = set()
        if whitelist_file:
            wl_raw = whitelist_file.read()
            if wl_raw[:3] == b'\xef\xbb\xbf':
                wl_raw = wl_raw[3:]
            elif wl_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                wl_raw = wl_raw.decode('utf-16').encode('utf-8')
            try:
                wl_content = wl_raw.decode("utf-8")
            except UnicodeDecodeError:
                wl_content = wl_raw.decode("latin-1")
            whitelist_terms = load_dictionary_terms(wl_content)
            st.success(f"🛡️ {len(whitelist_terms)} protected terms loaded")
            
            with st.expander("View protected terms"):
                for term in sorted(whitelist_terms):
                    st.write(f"- {term}")
        
        if mqxliff_files:
            st.markdown("---")
            if st.button("🚀 Process files", type="primary", use_container_width=True):
                process_files(
                    mqxliff_files, replacement_token, process_source, process_target,
                    use_safe_regex, use_regex, use_presidio, use_biomedical, use_proper_names,
                    use_dictionary, dictionary_terms, whitelist_terms
                )
    
    with tab2:
        if "previews" in st.session_state and st.session_state.previews:
            st.markdown("### Changes preview")
            
            col_search, col_show_junk = st.columns([3, 1])
            with col_search:
                search_term = st.text_input(
                    "🔍 Search in preview",
                    placeholder="Type to filter changes...",
                    key="preview_search"
                )
            with col_show_junk:
                show_junk_in_preview = st.checkbox(
                    "Show short",
                    value=False,
                    key="show_junk_preview",
                    help="Show/hide short segments in the preview (does not affect downloads)"
                )
            
            def calc_redaction_pct(text: str, token: str) -> float:
                if not text:
                    return 0
                words = text.split()
                if not words:
                    return 0
                redacted_count = sum(1 for w in words if token in w)
                return (redacted_count / len(words)) * 100
            
            replacement_token = st.session_state.get('replacement_token', '███')
            threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_enabled = st.session_state.get('exclude_modified_targets', False)
            exclude_source = st.session_state.get('exclude_source_too', False)
            
            if 'excluded_segments' not in st.session_state:
                st.session_state['excluded_segments'] = {}
            if 'no_anon_segments' not in st.session_state:
                st.session_state['no_anon_segments'] = {}
            if 'skip_junk_segments' not in st.session_state:
                st.session_state['skip_junk_segments'] = {}
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            
            total_matches = 0
            candidates_count = 0
            junk_segments_count = 0
            short_segments_count = 0
            
            for filename, file_previews in st.session_state.previews.items():
                filtered_previews = []
                for preview in file_previews:
                    is_junk = is_junk_segment(preview, min_words_junk)
                    segment_key_junk = f"{filename}_{preview['segment']}"
                    
                    if is_junk:
                        is_skipped = st.session_state.get(f"skipjunk_{segment_key_junk}", False)
                        if not is_skipped:
                            junk_segments_count += 1
                        if show_junk_in_preview:
                            preview['_is_junk'] = True
                            preview['_junk_skipped'] = is_skipped
                            filtered_previews.append(preview)
                        continue
                    
                    if not preview.get('changed', True):
                        continue
                    
                    has_token = (replacement_token in preview.get('source_after', '') or 
                                replacement_token in preview.get('target_after', ''))
                    if not has_token:
                        continue
                    
                    preview['_is_junk'] = False
                    if filter_short and segment_word_count(preview) < min_words:
                        short_segments_count += 1
                        continue
                    
                    if search_term:
                        search_lower = search_term.lower()
                        if (search_lower in preview['source_before'].lower() or
                            search_lower in preview['source_after'].lower() or
                            search_lower in preview['target_before'].lower() or
                            search_lower in preview['target_after'].lower()):
                            filtered_previews.append(preview)
                    else:
                        filtered_previews.append(preview)
                
                non_junk_count = sum(1 for p in filtered_previews if not p.get('_is_junk', False))
                junk_in_file = sum(1 for p in filtered_previews if p.get('_is_junk', False))
                
                if filtered_previews:
                    label_parts = []
                    if non_junk_count > 0:
                        label_parts.append(f"{non_junk_count} affected")
                    if junk_in_file > 0:
                        label_parts.append(f"{junk_in_file} short")
                    total_matches += non_junk_count
                    with st.expander(f"📄 {filename} ({', '.join(label_parts)} segments)", expanded=True):
                        for preview in filtered_previews:
                            preview_is_junk = preview.get('_is_junk', False)
                            segment_key = f"{filename}_{preview['segment']}"
                            
                            if preview_is_junk:
                                junk_skipped = preview.get('_junk_skipped', False)
                                col_header, col_skip_junk = st.columns([4, 1])
                                with col_header:
                                    if junk_skipped:
                                        st.markdown(
                                            f'<div style="background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#155724;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                            f'<br><small>EN: {preview["source_before"][:80] or "(empty)"} | ES: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        st.markdown(
                                            f'<div style="background:#e2e3e5;border:1px solid #6c757d;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#495057;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#6c757d;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Short</span>'
                                            f'<br><small>EN: {preview["source_before"][:80] or "(empty)"} | ES: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                with col_skip_junk:
                                    cb_key = f"skipjunk_{segment_key}"
                                    if cb_key not in st.session_state:
                                        st.session_state[cb_key] = False
                                    skip_junk = st.checkbox(
                                        "Skip",
                                        key=cb_key,
                                        help="Keep this segment (don't exclude as short)"
                                    )
                                    st.session_state['skip_junk_segments'][segment_key] = skip_junk
                                continue
                            
                            target_pct = calc_redaction_pct(preview.get('target_after', ''), replacement_token)
                            source_pct = calc_redaction_pct(preview.get('source_after', ''), replacement_token)
                            is_candidate = target_pct >= threshold
                            
                            if is_candidate:
                                candidates_count += 1
                            
                            has_changes = (preview['source_before'] != preview['source_after'] or 
                                         preview['target_before'] != preview['target_after'])
                            
                            col_header, col_no_anon, col_exclude = st.columns([3, 1, 1])
                            with col_header:
                                header_placeholder = st.empty()
                            
                            with col_no_anon:
                                if has_changes:
                                    no_anon_val = st.session_state['no_anon_segments'].get(segment_key, False)
                                    no_anon = st.checkbox(
                                        "Skip",
                                        value=no_anon_val,
                                        key=f"noanon_{segment_key}",
                                        help="Keep original text without anonymizing"
                                    )
                                    st.session_state['no_anon_segments'][segment_key] = no_anon
                            
                            with col_exclude:
                                if is_candidate and exclude_enabled:
                                    default_val = st.session_state['excluded_segments'].get(segment_key, True)
                                    exclude_this = st.checkbox(
                                        "Exclude TM",
                                        value=default_val,
                                        key=f"excl_{segment_key}"
                                    )
                                    st.session_state['excluded_segments'][segment_key] = exclude_this
                            
                            is_no_anon = st.session_state['no_anon_segments'].get(segment_key, False)
                            
                            segment_header = f"**Segment {preview['segment']}**"
                            if is_no_anon:
                                segment_header += ' <span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Not anonymized</span>'
                            elif is_candidate and exclude_enabled:
                                segment_header += f' <span class="exclude-badge">⚠️ {target_pct:.0f}% → exclusion</span>'
                            elif target_pct > 0:
                                segment_header += f' <small style="color:#666">({target_pct:.0f}%)</small>'
                            header_placeholder.markdown(segment_header, unsafe_allow_html=True)
                            
                            if preview['source_before'] != preview['source_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Source - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["source_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Source - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["source_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label_src = "**Source - After:**"
                                        show_excluded = is_candidate and exclude_enabled and exclude_source and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label_src = "**Source - After** _(will be excluded)_:"
                                        st.markdown(label_src)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["source_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            if preview['target_before'] != preview['target_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Target - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["target_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Target - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["target_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label = "**Target - After:**"
                                        show_excluded = is_candidate and exclude_enabled and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label = "**Target - After** _(will be excluded)_:"
                                        st.markdown(label)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["target_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            st.markdown("---")
            
            if filter_junk and junk_segments_count > 0:
                st.info(f"🗑️ {junk_segments_count} short segments excluded from TM (<{min_words_junk} words or only numbers/symbols)")
            
            if filter_short and short_segments_count > 0:
                st.info(f"✂️ {short_segments_count} short anonymized segments excluded from TM (less than {min_words} words)")
            
            if exclude_enabled and candidates_count > 0:
                excluded_count = sum(1 for v in st.session_state['excluded_segments'].values() if v)
                st.info(f"📊 {excluded_count} of {candidates_count} candidate segments will be excluded from TM (threshold: {threshold}%)")
            
            if search_term and total_matches == 0:
                st.warning(f"No changes found containing '{search_term}'")
        else:
            st.info("Process the files to see the changes preview")
    
    with tab3:
        if "results" in st.session_state and st.session_state.results:
            st.markdown("### Anonymization statistics")
            
            total_stats = {"safe_regex": 0, "regex_ct": 0, "presidio_pii": 0, "biomedical": 0, "proper_names": 0, "dictionary": 0}
            for stats in st.session_state.all_stats.values():
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)
            
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            render_stat_card("Safe Regex", total_stats["safe_regex"], col1, "stat-card-safe-regex")
            render_stat_card("Regex CT IDs", total_stats["regex_ct"], col2, "stat-card-regex-ct")
            render_stat_card("Presidio", total_stats["presidio_pii"], col3, "stat-card-presidio")
            render_stat_card("Biomedical", total_stats["biomedical"], col4, "stat-card-biomedical")
            render_stat_card("Proper Names", total_stats["proper_names"], col5, "stat-card-proper-names")
            render_stat_card("Dictionary", total_stats["dictionary"], col6, "stat-card-dictionary")
            
            st.markdown("---")
            st.markdown("### Download anonymized files")
            
            exclude_modified_targets = st.session_state.get('exclude_modified_targets', False)
            exclusion_threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_source_too = st.session_state.get('exclude_source_too', False)
            excluded_segments = st.session_state.get('excluded_segments', {})
            
            def get_junk_segment_keys(filename: str, file_previews: list, min_words_junk: int = 2) -> set:
                """Gets the keys of junk segments that should be cleared."""
                junk_keys = set()
                for preview in file_previews:
                    if is_junk_segment(preview, min_words_junk):
                        seg_key = f"{filename}_{preview['segment']}"
                        if not st.session_state.get(f"skipjunk_{seg_key}", False):
                            junk_keys.add(seg_key)
                return junk_keys
            
            def get_short_segment_keys(filename: str, file_previews: list, min_words: int) -> set:
                """Gets the keys of short segments that should be restored."""
                short_keys = set()
                for preview in file_previews:
                    if segment_word_count(preview) < min_words:
                        short_keys.add(f"{filename}_{preview['segment']}")
                return short_keys
            
            def apply_no_anon_segments(anon_content: bytes, orig_content: bytes, filename: str, no_anon_segs: dict) -> bytes:
                if not no_anon_segs:
                    return anon_content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    anon_tree = etree.fromstring(anon_content)
                    orig_tree = etree.fromstring(orig_content)
                    
                    if is_tmx:
                        anon_units = anon_tree.xpath('//tu')
                        orig_units = orig_tree.xpath('//tu')
                    else:
                        nsmap = anon_tree.nsmap
                        default_ns = nsmap.get(None, '')
                        
                        if default_ns:
                            ns = {'x': default_ns}
                            anon_units = anon_tree.xpath('//x:trans-unit', namespaces=ns)
                            orig_units = orig_tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            anon_units = anon_tree.xpath('//trans-unit')
                            orig_units = orig_tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, (anon_tu, orig_tu) in enumerate(zip(anon_units, orig_units)):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        should_restore_only = no_anon_segs.get(segment_key, False)
                        
                        if should_restore_only:
                            if is_tmx:
                                anon_tuvs = anon_tu.xpath('tuv')
                                orig_tuvs = orig_tu.xpath('tuv')
                                for anon_tuv, orig_tuv in zip(anon_tuvs, orig_tuvs):
                                    anon_segs = anon_tuv.xpath('seg')
                                    orig_segs = orig_tuv.xpath('seg')
                                    for anon_seg, orig_seg in zip(anon_segs, orig_segs):
                                        anon_seg.getparent().replace(anon_seg, orig_seg)
                            else:
                                if default_ns:
                                    anon_sources = anon_tu.xpath('.//x:source', namespaces=ns)
                                    orig_sources = orig_tu.xpath('.//x:source', namespaces=ns)
                                    anon_targets = anon_tu.xpath('.//x:target', namespaces=ns)
                                    orig_targets = orig_tu.xpath('.//x:target', namespaces=ns)
                                else:
                                    anon_sources = anon_tu.xpath('.//source')
                                    orig_sources = orig_tu.xpath('.//source')
                                    anon_targets = anon_tu.xpath('.//target')
                                    orig_targets = orig_tu.xpath('.//target')
                                
                                for anon_src, orig_src in zip(anon_sources, orig_sources):
                                    anon_src.getparent().replace(anon_src, orig_src)
                                
                                for anon_tgt, orig_tgt in zip(anon_targets, orig_targets):
                                    anon_tgt.getparent().replace(anon_tgt, orig_tgt)
                    
                    return etree.tostring(anon_tree, encoding='utf-8', xml_declaration=True)
                except:
                    return anon_content
            
            def _clear_element(elem):
                for child in list(elem):
                    elem.remove(child)
                elem.text = None
                elem.tail = None

            def prepare_download_content(content: bytes, filename: str, exclude_targets: bool, threshold: float, exclude_source: bool, excluded_segs: dict, short_segs: set = None, junk_segs: set = None) -> bytes:
                if short_segs is None:
                    short_segs = set()
                if junk_segs is None:
                    junk_segs = set()
                
                if not exclude_targets and not short_segs and not junk_segs:
                    return content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    tree = etree.fromstring(content)
                    
                    if is_tmx:
                        trans_units = tree.xpath('//tu')
                    else:
                        nsmap = tree.nsmap
                        default_ns = nsmap.get(None, '')
                        if default_ns:
                            ns = {'x': default_ns}
                            trans_units = tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            trans_units = tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, tu in enumerate(trans_units):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        is_short = segment_key in short_segs
                        is_junk = segment_key in junk_segs
                        
                        if is_short or is_junk:
                            if is_tmx:
                                for tuv in tu.xpath('tuv'):
                                    for seg in tuv.xpath('seg'):
                                        _clear_element(seg)
                            else:
                                if default_ns:
                                    targets = tu.xpath('.//x:target', namespaces=ns)
                                    sources = tu.xpath('.//x:source', namespaces=ns)
                                else:
                                    targets = tu.xpath('.//target')
                                    sources = tu.xpath('.//source')
                                for elem in sources + targets:
                                    _clear_element(elem)
                            continue
                        
                        if not exclude_targets:
                            continue
                        
                        should_exclude_this = excluded_segs.get(segment_key, None)
                        if should_exclude_this is False:
                            continue
                        
                        if is_tmx:
                            target_tuvs = []
                            source_tuvs = []
                            for tuv in tu.xpath('tuv'):
                                tuv_lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                                if tuv_lang.lower().startswith("es"):
                                    target_tuvs.extend(tuv.xpath('seg'))
                                elif tuv_lang.lower().startswith("en") and exclude_source:
                                    source_tuvs.extend(tuv.xpath('seg'))
                            
                            for seg in target_tuvs:
                                text_content = ''.join(seg.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    if redacted_pct >= threshold:
                                        _clear_element(seg)
                                        if exclude_source:
                                            for src_seg in source_tuvs:
                                                _clear_element(src_seg)
                        else:
                            if default_ns:
                                targets = tu.xpath('.//x:target', namespaces=ns)
                                sources = tu.xpath('.//x:source', namespaces=ns) if exclude_source else []
                            else:
                                targets = tu.xpath('.//target')
                                sources = tu.xpath('.//source') if exclude_source else []
                            
                            for target in targets:
                                text_content = ''.join(target.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    
                                    if redacted_pct >= threshold:
                                        _clear_element(target)
                                        
                                        if exclude_source:
                                            for source in sources:
                                                _clear_element(source)
                    
                    return etree.tostring(tree, encoding='utf-8', xml_declaration=True)
                except:
                    return content
            
            no_anon_segments = st.session_state.get('no_anon_segments', {})
            originals = st.session_state.get('originals', {})
            
            no_anon_count = sum(1 for v in no_anon_segments.values() if v)
            if no_anon_count > 0:
                st.markdown(f"""
                <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>✓ {no_anon_count} segments will keep their original text</strong><br>
                    <small>Marked as "Skip" in the Preview tab</small>
                </div>
                """, unsafe_allow_html=True)
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            previews = st.session_state.get('previews', {})
            
            dedup_tmx = st.session_state.get('dedup_tmx', True)
            dedup_threshold = st.session_state.get('dedup_threshold', 100)
            
            clean_tmx_data, valid_segments, no_anon_in_tmx, excluded_ids, dedup_count, dedup_details = generate_clean_tmx(
                previews=previews,
                results=st.session_state.results,
                originals=originals,
                filter_junk=filter_junk,
                min_words_junk=min_words_junk,
                filter_short=filter_short,
                min_words=min_words,
                exclude_modified=exclude_modified_targets,
                exclusion_threshold=exclusion_threshold,
                excluded_segments=excluded_segments,
                no_anon_segments=no_anon_segments,
                dedup_tmx=dedup_tmx,
                dedup_threshold=dedup_threshold
            )
            
            st.session_state['dedup_details'] = dedup_details
            
            total_segs = sum(len(fp) for fp in previews.values())
            empty_total = total_segs - valid_segments - no_anon_in_tmx
            if empty_total > 0:
                st.markdown(f"""
                <div style="background-color: #e2e3e5; border: 1px solid #6c757d; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>🗑️ {empty_total} segments excluded from clean TMX</strong> <small>(of {total_segs} total)</small><br>
                    <small>Excluded by filters (short segments, short anon. segments, heavily anonymized, duplicates). Not included in clean TMX</small>
                </div>
                """, unsafe_allow_html=True)
            
            col_mqxliff, col_tmx_clean, col_excel = st.columns(3)
            
            with col_mqxliff:
                if len(st.session_state.results) == 1:
                    filename, content = list(st.session_state.results.items())[0]
                    orig_content = originals.get(filename, content)
                    junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                    short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                    content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                    download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                    file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                    redacted_filename = f"Redacted_{date.today().isoformat()}.{file_ext}"
                    st.download_button(
                        label=f"📥 Download {file_ext.upper()}",
                        data=download_content,
                        file_name=redacted_filename,
                        mime="application/xml",
                        use_container_width=True
                    )
                else:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for idx, (filename, content) in enumerate(st.session_state.results.items()):
                            orig_content = originals.get(filename, content)
                            junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                            short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                            content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                            download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                            file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                            zip_entry_name = f"Redacted_{idx + 1}_{date.today().isoformat()}.{file_ext}"
                            zf.writestr(zip_entry_name, download_content)
                    
                    st.download_button(
                        label=f"📥 Download ZIP ({len(st.session_state.results)} files)",
                        data=zip_buffer.getvalue(),
                        file_name=f"Redacted_{date.today().isoformat()}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            
            with col_tmx_clean:
                st.download_button(
                    label="📥 Download clean TMX",
                    data=clean_tmx_data,
                    file_name=f"Redacted_TM_{date.today().isoformat()}.tmx",
                    mime="application/xml",
                    use_container_width=True,
                    help="TMX without empty, excluded or filtered segments"
                )
            
            with col_excel:
                excel_data = generate_changes_excel(dedup_details=st.session_state.get('dedup_details', []))
                st.download_button(
                    label="📊 Download changes report",
                    data=excel_data,
                    file_name="anonymization_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("Process the files to download them")


def process_files(files, replacement_token, process_source, process_target,
                  use_safe_regex, use_regex, use_presidio, use_biomedical, use_proper_names,
                  use_dictionary, dictionary_terms, whitelist_terms=None):
    
    anonymizer = MQXLIFFAnonymizer(replacement_token=replacement_token)
    
    results = {}
    originals = {}
    all_stats = {}
    previews = {}
    
    progress_bar = st.progress(0, text="Processing files...")
    
    for i, file in enumerate(files):
        file_label = file.name if len(file.name) <= 40 else file.name[:37] + "..."
        
        def make_progress_cb(file_idx, total_files, fname):
            last_pct = [-1]
            def cb(current, total):
                pct = int((current / total) * 100) if total > 0 else 0
                if pct == last_pct[0] and pct < 100:
                    return
                last_pct[0] = pct
                file_base = file_idx / total_files
                file_share = 1.0 / total_files
                segment_pct = current / total if total > 0 else 0
                overall = file_base + file_share * segment_pct
                progress_bar.progress(min(overall, 1.0), text=f"Processing {fname}... {pct}% ({current}/{total} segments)")
                time.sleep(0.01)
            return cb
        
        progress_cb = make_progress_cb(i, len(files), file_label)
        progress_cb(0, 1)
        
        try:
            content = file.read()
            originals[file.name] = content
            
            is_tmx = file.name.lower().endswith(".tmx")
            
            if is_tmx:
                result_xml, stats, file_previews = anonymizer.anonymize_tmx(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_presidio=use_presidio,
                    use_biomedical=use_biomedical,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            else:
                result_xml, stats, file_previews = anonymizer.anonymize_mqxliff(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_presidio=use_presidio,
                    use_biomedical=use_biomedical,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            
            results[file.name] = result_xml
            all_stats[file.name] = stats
            previews[file.name] = file_previews
            
        except Exception as e:
            st.error(f"Error processing {file.name}: {str(e)}")
    
    progress_bar.empty()
    
    st.session_state.results = results
    st.session_state.originals = originals
    st.session_state.all_stats = all_stats
    st.session_state.previews = previews
    
    st.success(f"✅ Successfully processed {len(results)} file(s)")
    st.info("Go to the **Preview** and **Download** tabs to see the results")


if __name__ == "__main__":
    main()
